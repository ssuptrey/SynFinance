"""
Tests for Excel Dashboard Generator.

Author: SynFinance Development Team
Date: November 2, 2025
"""

import pytest
import pandas as pd
import numpy as np
from pathlib import Path
import tempfile
from openpyxl import load_workbook

from src.reporting.excel_generator import ExcelDashboardGenerator


class TestExcelDashboardGenerator:
    """Test suite for Excel dashboard generation."""
    
    @pytest.fixture
    def sample_data(self):
        """Create sample transaction data."""
        np.random.seed(42)
        return pd.DataFrame({
            'Transaction_ID': range(1, 101),
            'Transaction_Amount': np.random.lognormal(8, 1.5, 100),
            'Category': np.random.choice(['Groceries', 'Dining', 'Shopping'], 100),
            'Customer_ID': np.random.randint(1, 20, 100),
            'Fraud_Type': [None] * 95 + ['Card Cloning'] * 5
        })
    
    @pytest.fixture
    def generator(self):
        """Create ExcelDashboardGenerator instance."""
        return ExcelDashboardGenerator()
    
    def test_generator_initialization(self, generator):
        """Test ExcelDashboardGenerator initializes correctly."""
        assert generator is not None
        assert generator.colors is not None
        assert 'header_bg' in generator.colors
    
    def test_create_dashboard_workbook(self, generator, sample_data):
        """Test creating complete dashboard workbook."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test_dashboard.xlsx'
            
            result_path = generator.create_dashboard_workbook(
                data=sample_data,
                output_path=output_path,
                include_charts=True
            )
            
            assert result_path.exists()
            assert result_path.stat().st_size > 0
            
            # Load and verify workbook
            wb = load_workbook(output_path)
            assert 'Summary Dashboard' in wb.sheetnames
            assert 'Transaction Data' in wb.sheetnames
            assert 'Statistical Analysis' in wb.sheetnames
            assert 'Fraud Analysis' in wb.sheetnames
            wb.close()
    
    def test_add_summary_sheet(self, generator, sample_data):
        """Test adding summary sheet."""
        from openpyxl import Workbook
        wb = Workbook()
        
        ws = generator.add_summary_sheet(wb, sample_data)
        
        assert ws is not None
        assert ws['A1'].value == 'SynFinance Executive Dashboard'
        assert ws['B3'].value == 100  # Total transactions
    
    def test_add_data_sheet(self, generator, sample_data):
        """Test adding data sheet."""
        from openpyxl import Workbook
        wb = Workbook()
        
        ws = generator.add_data_sheet(wb, sample_data)
        
        assert ws is not None
        # Check headers
        assert ws.cell(1, 1).value in sample_data.columns
        # Check frozen panes
        assert ws.freeze_panes is not None
    
    def test_add_statistics_sheet(self, generator, sample_data):
        """Test adding statistics sheet."""
        from openpyxl import Workbook
        wb = Workbook()
        
        ws = generator.add_statistics_sheet(wb, sample_data)
        
        assert ws is not None
        assert ws['A1'].value == 'Statistical Summary'
        # Check for statistical headers
        assert 'Field' in [ws.cell(3, i).value for i in range(1, 7)]
    
    def test_add_fraud_analysis_sheet(self, generator, sample_data):
        """Test adding fraud analysis sheet."""
        from openpyxl import Workbook
        wb = Workbook()
        
        ws = generator.add_fraud_analysis_sheet(wb, sample_data)
        
        assert ws is not None
        assert 'Fraud Detection Analysis' in ws['A1'].value
        assert ws.cell(3, 1).value == 'Total Fraudulent Transactions'
        assert ws.cell(3, 2).value == 5  # 5 fraud transactions in sample
    
    def test_add_charts_sheet(self, generator, sample_data):
        """Test adding charts sheet."""
        from openpyxl import Workbook
        wb = Workbook()
        
        ws = generator.add_charts_sheet(wb, sample_data)
        
        assert ws is not None
        assert 'Visual Analytics Dashboard' in ws['A1'].value
    
    def test_workbook_without_fraud_data(self, generator):
        """Test workbook creation without fraud columns."""
        data = pd.DataFrame({
            'Transaction_Amount': [100, 200, 300],
            'Category': ['A', 'B', 'C']
        })
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test_no_fraud.xlsx'
            
            result_path = generator.create_dashboard_workbook(
                data=data,
                output_path=output_path,
                include_charts=False
            )
            
            assert result_path.exists()
            wb = load_workbook(output_path)
            # Should not have fraud analysis sheet
            assert 'Fraud Analysis' not in wb.sheetnames
            wb.close()
    
    def test_workbook_with_no_charts(self, generator, sample_data):
        """Test workbook creation without charts."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test_no_charts.xlsx'
            
            result_path = generator.create_dashboard_workbook(
                data=sample_data,
                output_path=output_path,
                include_charts=False
            )
            
            assert result_path.exists()
            wb = load_workbook(output_path)
            assert 'Charts & Visualizations' not in wb.sheetnames
            wb.close()
    
    def test_large_dataset(self, generator):
        """Test Excel generation with large dataset."""
        large_data = pd.DataFrame({
            'Transaction_Amount': np.random.lognormal(8, 1.5, 5000),
            'Category': np.random.choice(['A', 'B', 'C'], 5000)
        })
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test_large.xlsx'
            
            result_path = generator.create_dashboard_workbook(
                data=large_data,
                output_path=output_path
            )
            
            assert result_path.exists()
            assert result_path.stat().st_size > 0


class TestExcelEdgeCases:
    """Test edge cases for Excel generation."""
    
    @pytest.fixture
    def generator(self):
        return ExcelDashboardGenerator()
    
    def test_empty_dataframe(self, generator):
        """Test handling of empty DataFrame."""
        empty_df = pd.DataFrame()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test_empty.xlsx'
            
            result_path = generator.create_dashboard_workbook(
                data=empty_df,
                output_path=output_path
            )
            
            assert result_path.exists()
    
    def test_single_row_dataframe(self, generator):
        """Test handling of single row DataFrame."""
        single_row = pd.DataFrame({'Amount': [100]})
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test_single.xlsx'
            
            result_path = generator.create_dashboard_workbook(
                data=single_row,
                output_path=output_path
            )
            
            assert result_path.exists()
    
    def test_special_characters_in_data(self, generator):
        """Test handling of special characters."""
        data = pd.DataFrame({
            'Category': ['Test₹', 'Test€', 'Test¥'],
            'Amount': [100, 200, 300]
        })
        
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test_special.xlsx'
            
            result_path = generator.create_dashboard_workbook(
                data=data,
                output_path=output_path
            )
            
            assert result_path.exists()


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
